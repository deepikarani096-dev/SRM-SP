"""
Abstract scraper for papers CSV.
Uses the free Crossref API (no API key required).
Falls back to Semantic Scholar and OpenAlex for missing abstracts.

Usage:
    python scrape_abstracts.py --input papers.csv --output papers_with_abstracts.xlsx
    python scrape_abstracts.py --input papers.csv --output papers_with_abstracts.xlsx --limit 100
    python scrape_abstracts.py --input papers.csv --output papers_with_abstracts.xlsx --resume
"""

import argparse
import csv
import time
import json
import sys
import re
import urllib.request
import urllib.parse
import urllib.error
from pathlib import Path

CACHE_FILE = Path("abstract_cache.json")


def get_json(url, timeout=15):
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "AbstractScraper/1.0 (research tool; mailto:research@example.com)",
            "Accept": "application/json",
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return json.loads(r.read().decode())
    except Exception:
        return None


def strip_html(text):
    if not text:
        return text
    return re.sub(r"<[^>]+>", "", text).strip()


def fetch_crossref(doi):
    url = f"https://api.crossref.org/works/{urllib.parse.quote(doi, safe='')}"
    data = get_json(url)
    if data and data.get("status") == "ok":
        abstract = data["message"].get("abstract", "")
        return strip_html(abstract) if abstract else None
    return None


def fetch_semantic_scholar(doi):
    url = f"https://api.semanticscholar.org/graph/v1/paper/DOI:{urllib.parse.quote(doi, safe='')}?fields=abstract"
    data = get_json(url)
    if data and data.get("abstract"):
        return data["abstract"].strip()
    return None


def fetch_openalex(doi):
    url = f"https://api.openalex.org/works/https://doi.org/{urllib.parse.quote(doi, safe='')}"
    data = get_json(url)
    if data and data.get("abstract_inverted_index"):
        inv = data["abstract_inverted_index"]
        positions = {}
        for word, idxs in inv.items():
            for idx in idxs:
                positions[idx] = word
        abstract = " ".join(positions[k] for k in sorted(positions))
        return abstract.strip() if abstract.strip() else None
    return None


def get_abstract(doi, delay=0.5):
    if not doi or not doi.strip():
        return "No DOI provided"

    doi = doi.strip()

    abstract = fetch_crossref(doi)
    if abstract:
        return abstract
    time.sleep(delay)

    abstract = fetch_semantic_scholar(doi)
    if abstract:
        return abstract
    time.sleep(delay)

    abstract = fetch_openalex(doi)
    if abstract:
        return abstract

    return "Abstract not found"


def load_cache():
    if CACHE_FILE.exists():
        with open(CACHE_FILE) as f:
            return json.load(f)
    return {}


def save_cache(cache):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f)


def main():
    parser = argparse.ArgumentParser(description="Scrape paper abstracts via DOI.")
    parser.add_argument("--input",  default="papers.csv", help="Input CSV file")
    parser.add_argument("--output", default="papers_with_abstracts.xlsx", help="Output Excel file")
    parser.add_argument("--limit",  type=int, default=None, help="Max rows to process")
    parser.add_argument("--delay",  type=float, default=0.5, help="Seconds between API requests")
    parser.add_argument("--resume", action="store_true", help="Skip DOIs already in cache")
    args = parser.parse_args()

    with open(args.input, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    if args.limit:
        rows = rows[:args.limit]

    cache = load_cache() if args.resume else {}
    total = len(rows)

    print(f"Processing {total} papers...")
    print(f"Sources: Crossref -> Semantic Scholar -> OpenAlex\n")

    found = 0
    not_found = 0

    for i, row in enumerate(rows, 1):
        doi = row.get("doi", "").strip()

        if doi and doi in cache:
            row["abstract"] = cache[doi]
            status = "cached"
        else:
            abstract = get_abstract(doi, delay=args.delay)
            row["abstract"] = abstract
            if doi:
                cache[doi] = abstract
            if abstract and abstract not in ("Abstract not found", "No DOI provided"):
                found += 1
                status = "found"
            else:
                not_found += 1
                status = "not found"
            if i % 25 == 0:
                save_cache(cache)

        title_preview = row.get("title", "")[:50]
        print(f"[{i:4d}/{total}] {status:<10} | {doi[:30] if doi else 'no doi':<30} | {title_preview}")
        sys.stdout.flush()

    save_cache(cache)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Papers"

        cols = list(rows[0].keys())
        header_fill = PatternFill("solid", start_color="2B5F8E", end_color="2B5F8E")
        header_font = Font(bold=True, color="FFFFFF", name="Arial")

        for col_idx, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(rows, 2):
            for col_idx, col in enumerate(cols, 1):
                val = row.get(col, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=(col == "abstract"))
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", start_color="F2F7FC", end_color="F2F7FC")

        widths = {"doi": 28, "title": 50, "publication_name": 35, "date": 14, "abstract": 80}
        for col_idx, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col, 20)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        wb.save(args.output)
        print(f"\nSaved to: {args.output}")

    except ImportError:
        csv_path = args.output.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        print(f"\nSaved as CSV (openpyxl not installed): {csv_path}")

    print(f"\nSummary:")
    print(f"  Total papers    : {total}")
    print(f"  Abstracts found : {found}")
    print(f"  Not found       : {not_found}")


if __name__ == "__main__":
    main()